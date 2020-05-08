from flask import Flask, render_template, flash, redirect, url_for, session, request,send_file
from werkzeug.debug import DebuggedApplication
#from data import Articles
import os
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
from werkzeug.middleware.shared_data import SharedDataMiddleware
from flask import send_from_directory
from wtforms import Form, StringField, TextAreaField, PasswordField, validators
from wtforms.validators import Email ,ValidationError
from werkzeug.utils import secure_filename
from functools import wraps
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from datetime import datetime
from flask_login import UserMixin
from flask_login import login_user, current_user, logout_user, login_required,LoginManager
import flask_excel as excel

app  = Flask(__name__ , template_folder='templetes')

UPLOAD_FOLDER='C:\\Users\\shivam.a.shrivastava\\PycharmProjects\\TESt'

sqldb = SQLAlchemy()
app.config["SQLALCHEMY_DATABASE_URI"] = "mysql+pymysql://root:mysql2020@localhost:3306/tax"
app.config['SQLALCHEMY_ECHO'] = True
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_POOL_TIMEOUT'] = 20
app.config['SQLALCHEMY_POOL_RECYCLE'] = 299
app.config['SQLALCHEMY_MAX_OVERFLOW'] = 20
app.config['SQLALCHEMY_POOL_SIZE'] = 10
app.config['UPLOAD_FOLDER']=UPLOAD_FOLDER
db = SQLAlchemy(app)
db.init_app(app)
bcrypt=Bcrypt(app)

app.add_url_rule('/uploads/<filename>', 'uploaded_file',build_only=True)
app.wsgi_app = SharedDataMiddleware(app.wsgi_app, {'/uploads':  app.config['UPLOAD_FOLDER']})

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'info'


@login_manager.user_loader
def load_user(user_id):
    return SHI.query.get(int(user_id))


class SHI(db.Model,UserMixin):
    id=db.Column(db.Integer,primary_key=True,autoincrement=True)
    name= db.Column(db.String(80), nullable=False)
    username = db.Column(db.String(80),  nullable=False)
    email = db.Column(db.String(120), unique=True , nullable=False)
    password = db.Column(db.String(80),nullable=False)
    teams =db.Column(db.String(120),nullable=True)

    def __repr__(self):
        return '<User %r>' % self.username

class POST(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    title= db.Column(db.String(80),nullable=False)
    body = db.Column(db.Text, nullable=False)
    Date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    author= db.Column(db.String(80),nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('SHI.id'), nullable=True)

    def __repr__(self):
        return f"Post('{self.title}')"
class Filesav(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name=db.Column(db.String(300))
    data=db.Column(db.LargeBinary(length=(2**32)-1))
    Date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    user_id=db.Column(db.Integer, db.ForeignKey('SHI.id'), nullable=True)


class Training(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    username= db.Column(db.String(80), nullable=False)
    Total=db.Column(db.Integer, autoincrement=True)

    completed=db.Column(db.Integer,  autoincrement=True)
    left=db.Column(db.Integer,  autoincrement=True)
    Date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    Training_name = db.Column(db.String(80), nullable=False)
    user_id=db.Column(db.Integer, db.ForeignKey('SHI.id'), nullable=True)


#db.create_all()
#Articles = Articles()
# Index
@app.route('/')
def index():
    return render_template('home.html')


# About
@app.route('/about')
def about():
    return render_template('about.html')


# Articles
@app.route('/articles')
def articles():
    page = request.args.get('page', 1, type=int)
    articles=POST.query.all().paginate(page=page, per_page=2)
    return render_template('articles.html', articles=articles)

#Single Article
@app.route('/article/<string:id>/')
def article(id):
    article=POST.query.filter_by(id=int(id)).first()
    return render_template('article.html', article=article)


# Register Form Class
class RegisterForm(Form):
    name = StringField('Name', [validators.Length(min=1, max=50)])
    username = StringField('Username', [validators.Length(min=4, max=25)])
    email = StringField('Email', [validators.Length(min=6, max=50),Email()])
    password = PasswordField('Password', [
        validators.DataRequired(),
        validators.EqualTo('confirm', message='Passwords do not match')
    ])
    confirm = PasswordField('Confirm Password')

    def validate_username(self,username):
        user = SHI.query.filter_by(username=username.data).first()
        #print(user.username)
        print(username.data + "data")
        if user:
            raise ValidationError('That username is taken. Please choose a different one.')

    def validate_email(self, email):
        user = SHI.query.filter_by(email=email.data).first()
        if user:
            raise ValidationError('That email is taken. Please choose a different one.')


# User Register
@app.route('/register', methods=['GET', 'POST'])
def register():
    form = RegisterForm(request.form)
    if request.method == 'POST' and form.validate():
        #password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        me=SHI(name=form.name.data,email=form.email.data,username=form.username.data,password=form.password.data,teams=request.form['teams'])
        db.session.add(me)
        db.session.commit()

        flash(f'You are now registered {form.name.data} and can log in', 'success')

        return redirect(url_for('login'))
    return render_template('register.html', form=form)

# User login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password_candidate = request.form['password']
        user = SHI.query.filter_by(username=username).first()

        if user and password_candidate == user.password:
            login_user(user)
            #next_page = request.args.get('next')
            return  redirect(url_for('dashboard'))
        else:
            flash('Please enter correct credentials', 'danger')
            return redirect(url_for('login'))
    return render_template('login.html')


@app.route('/UpdateYourProfile/<string:id>',methods=['GET','POST'])
@login_required
def updateyourprofile(id):
    details=SHI.query.filter_by(id=id).first()

    form = RegisterForm(request.form)

    if request.method =="GET":
        form.username.data=details.username
        form.name.data=details.name
        form.email.data=details.email

    elif request.method == "POST" and form.validate():
        details.username=form.username.data
        details.name=form.name.data
        details.email=form.email.data
        details.password=form.password.data
        details.teams=request.form['teams']
        db.session.commit()

        flash('Your Profile is updated login again')
        return redirect(url_for('logout'))
    return render_template('updateyourprofile.html' ,form=form)

# Logout
@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash(f'You are now logged out', 'success')
    return redirect(url_for('login'))

# Dashboard
@app.route('/dashboard')
@login_required
def dashboard():
    articles = POST.query.filter_by(user_id=current_user.id).all()
    le=len(articles)
    print(le)
    return render_template('dashboard.html',articles=articles,le=le)


# Article Form Class
class ArticleForm(Form):
    title = StringField('Title', [validators.Length(min=1, max=200)])
    #author = StringField('Author', [validators.Length(min=1, max=200),validators.DataRequired()])
    body = TextAreaField('Body', [validators.Length(min=10)])

# Add Article
@app.route('/add_article', methods=['GET', 'POST'])
@login_required
def add_article():
    form = ArticleForm(request.form)
    if request.method == 'POST' and form.validate():
        ad=POST(title=form.title.data,body=form.body.data,author=current_user.username,user_id=current_user.id)
        db.session.add(ad)
        db.session.commit()
        flash(f'Article Created {form.title.data}', 'success')

        return redirect(url_for('dashboard'))

    return render_template('add_article.html', form=form)

#type of file allowed
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'}
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
###############################################################################
#Add files
@app.route('/Add_Files',methods=['GET','POST'])
@login_required
def add_files():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash(f'No file part', 'danger')
            return redirect(url_for('dashboard'))
        fil = request.files['file']
        # if user does not select file, browser also
        # submit an empty part without filename
        if fil.filename == '':
            flash(f'No selected file', 'danger')
            return redirect(url_for('dashboard'))
        if fil:
            sav = Filesav(name=fil.filename, data=fil.read(), user_id=current_user.id)
            db.session.add(sav)
            db.session.commit()
            flash(f' your file {fil.filename} is saved succesfully', 'success')
            return redirect(url_for('dashboard'))
    return render_template('Add_files.html')

@app.route('/Download_Files')
@login_required
def download():
    view=Filesav.query.all()
    return render_template('downloadfiles.html',view=view)

    return render_template('downloadfiles.html')


@app.route('/downloadfiles/<string:id>',methods=['GET','POST'])
@login_required
def downloadfile(id):
    fil_data=Filesav.query.filter_by(id=id).first()
    return send_file(BytesIO(fil_data.data),attachment_filename=fil_data.name ,as_attachment=True)

@app.route('/delete_uploaded_Files',methods=['GET','POST'])
@login_required
def deletefile():
    filez=Filesav.query.filter_by(user_id=current_user.id).all()
    return render_template('deletefiles.html',filez=filez)

@app.route('/deletefile/<string:id>', methods=['GET','POST'])
@login_required
def deletefiledb(id):
    fildel=Filesav.query.filter_by(id=id).first()
    print(fildel.name)
    db.session.delete(fildel)
    db.session.commit()
    flash(f'your file is deleted successfully','success')
    return redirect(url_for('dashboard'))

#########################################################################

"""@app.route('/uploads/<filename>')
#@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'],filename)
"""# Edit Article
@app.route('/edit_article/<string:id>', methods=['GET', 'POST'])
@login_required
def edit_article(id):
    # Create cursor
    post = POST.query.filter_by(id=int(id)).first()


    form = ArticleForm(request.form)

    # Populate article form fields
    if request.method =="GET":
        form.title.data = post.title
        form.body.data = post.body

    elif request.method == 'POST' :
        post.title= form.title.data
        print(form.title.data)
        post.body = form.body.data
        print(form.body.data)
        #POST.author =request.form['author']
        db.session.commit()

        flash('Article Updated', 'success')

        return redirect(url_for('dashboard'))

    return render_template('edit_article.html', form=form)

# Delete Article
@app.route('/delete_article/<string:id>', methods=['POST'])
@login_required
def delete_article(id):
    # Create cursor
    """
    d=POST.query.get_or_404(int(id))
    db.session.delete(d)
    db.session.commit()"""
    me=POST.query.filter_by(id=id).first()
    db.session.delete(me)
    db.session.commit()

    flash(f'Article Deleted ', 'success')

    return  redirect(url_for('dashboard'))


###############################################training part####################
class Traning(Form):
    Total =  StringField('Total', [validators.length(min=1, max=10)])
    completed=  StringField('Completed',[validators.length(min=1, max=10)])

@app.route('/traning_update',methods=['GET','POST'])
@login_required
def training_update():
    form = Traning(request.form)
    if request.method == 'POST' and form.validate():
        left=(int(form.Total.data) -int(form.completed.data))
        username=current_user.username
        user_id=current_user.id
        Training_name = request.form['Training']
        add_tran=Training(username=username,Total=form.Total.data,completed=form.completed.data,left=left,user_id=user_id,Training_name=Training_name)
        db.session.add(add_tran)
        db.session.commit()
        flash(f'Your training details added successfully','success')
        return redirect(url_for('dashboard'))

    return render_template('training_updates.html',form=form)

@app.route("/export", methods=['GET','POST'])
@login_required
def export_records():
    if request.method == 'POST':
        name=request.form['Training']
        filename=datetime.now().strftime("%Y%m%d%H%M%S%p-") + current_user.username + '.xlsx'
        read=Training.query.filter_by(Training_name=name).all()
        wb = Workbook(write_only=True)
        city_ws = wb.create_sheet(f'Training')
        # write header
        city_ws.append(["username", "Total","completed","left","Training_name"])
        for city in read:
            username= city.username
            Total = city.Total
            completed=city.completed
            left=city.left
            Training_name=city.Training_name
            city_ws.append([username, Total,completed,left,Training_name])

        wb.save(filename)
        return send_from_directory(app.config['UPLOAD_FOLDER'],filename, as_attachment=True)

    return render_template('export.html')

@app.route("/VIEW", methods=['GET','POST'])
@login_required
def view_records():
    if request.method == 'POST':
        name=request.form['Training']
        filename=datetime.now().strftime("%Y%m%d%H%M%S%p-") + current_user.username + '.xlsx'
        read=Training.query.filter_by(Training_name=name).all()
        wb = Workbook(write_only=True)
        city_ws = wb.create_sheet(f'Training')
        # write header
        city_ws.append(["username", "Total","completed","left","Training_name"])
        for city in read:
            username= city.username
            Total = city.Total
            completed=city.completed
            left=city.left
            Training_name=city.Training_name
            city_ws.append([username, Total,completed,left,Training_name])

        wb.save(filename)
        df=pd.read_excel(filename,sheet_name='Training')
        return df.to_html()
    return render_template('view.html')

@app.route('/UPDATE_TRAIN',methods=['GET','POST'])
@login_required
def update_training():
    tview=Training.query.filter_by(user_id=current_user.id).all()
    return render_template('update_train.html',tview=tview)

@app.route('/UPDATE-TRAIN/<string:id>',methods=['GET','POST'])
@login_required
def update_tran_p(id):
    pas=Training.query.filter_by(id=id).first()

    form = Traning(request.form)
    if request.method =="GET":
        form.Total.data = pas.Total
        form.completed.data = pas.completed
    elif request.method == 'POST':
        pas.Total=form.Total.data
        pas.completed=form.completed.data
        pas.left=(int(form.Total.data)-int(form.completed.data))
        db.session.commit()
        flash(f'your training details updated successfully')
        return redirect(url_for('dashboard'))


    return render_template('update_html.html',form=form)






if __name__ == '__main__':
    app.secret_key='secret123'
    app.run(host='192.168.1.104',port=5000)